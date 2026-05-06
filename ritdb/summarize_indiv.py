import openpyxl
from collections import defaultdict

INPUT_FILE = "final_indiv.xlsx"
OUTPUT_FILE = "fin_indiv_sum.xlsx"

wb_in = openpyxl.load_workbook(INPUT_FILE)
sheet_names = wb_in.sheetnames  # ['Game1', 'Game2', 'Game3', 'Sheet4']

# counts[game_label][player] = {'15s': int, '10s': int}
game_data = {}

for i, name in enumerate(sheet_names, start=1):
    label = f"Game{i}"
    ws = wb_in[name]
    counts = defaultdict(lambda: {"15s": 0, "10s": 0})
    for row in ws.iter_rows(min_row=1, values_only=True):
        player, pts = row[0], row[1]
        if player is None:
            continue
        if pts == 15:
            counts[player]["15s"] += 1
        elif pts == 10:
            counts[player]["10s"] += 1
    game_data[label] = counts


def make_rows(counts):
    rows = []
    for player, c in counts.items():
        pts = 15 * c["15s"] + 10 * c["10s"]
        statline = f"{c['15s']}/{c['10s']}"
        rows.append((player, statline, pts))
    rows.sort(key=lambda r: -r[2])
    return rows


wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)  # remove default sheet

for label in ["Game1", "Game2", "Game3", "Game4"]:
    ws = wb_out.create_sheet(label)
    ws.append(["Player Name", "Statline", "Points"])
    for row in make_rows(game_data[label]):
        ws.append(list(row))

# Overall summary: sum 10s and 15s across all games per player
overall = defaultdict(lambda: {"15s": 0, "10s": 0})
for label in ["Game1", "Game2", "Game3", "Game4"]:
    for player, c in game_data[label].items():
        overall[player]["15s"] += c["15s"]
        overall[player]["10s"] += c["10s"]

ws_sum = wb_out.create_sheet("Overall")
ws_sum.append(["Player Name", "Statline", "Points"])
for row in make_rows(overall):
    ws_sum.append(list(row))

wb_out.save(OUTPUT_FILE)
print(f"Saved {OUTPUT_FILE}")
